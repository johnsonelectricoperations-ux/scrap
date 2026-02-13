"""
폐기불량 관리시스템 - Python Flask 버전
내부 네트워크용 (인터넷 불필요)
Excel + SQLite DB 이중 저장 (Excel 백업 + DB 통계/분석)
"""

from flask import Flask, render_template, request, jsonify
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime
import sqlite3
import os
import random

app = Flask(__name__)

# 설정
EXCEL_FILE = 'scrap_data.xlsx'
DB_FILE = 'scrap_data.db'
ADMIN_PASSWORD = 'admin1234'

# Excel 시트 구조 정의
SHEET_STRUCTURES = {
    'Depart': ['부서'],
    'Process': ['Part', '공정'],
    'machine': ['Part', '공정', '설비명'],
    'person': ['Part', '공정', '부서', '폐기자'],
    '1Part_TMNO': ['TM-NO', '품명', '단위중량', '성형', '소결', '후처리'],
    '2Part_TMNO': ['TM-NO', '품명', '단위중량', '성형', '소결', '후처리'],
    'scrap_name': ['폐기사유'],
    'Data': ['ID', '날짜', 'Part', '부서', '공정', '설비명', '폐기자', 'TM-NO', '품명', '폐기사유', '수량', '중량(kg)', '비고']
}


# ==================== Excel 함수 ====================

def init_excel_file():
    """Excel 파일 초기화 (없으면 생성)"""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        default_sheet = wb.active

        for sheet_name, headers in SHEET_STRUCTURES.items():
            ws = wb.create_sheet(title=sheet_name)
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        wb.save(EXCEL_FILE)
        print(f"Excel 파일 생성됨: {EXCEL_FILE}")
    return True


def get_workbook():
    """Excel 워크북 가져오기"""
    init_excel_file()
    return load_workbook(EXCEL_FILE)


def save_workbook(wb):
    """Excel 워크북 저장"""
    wb.save(EXCEL_FILE)


# ==================== SQLite DB 함수 ====================

def get_db():
    """DB 연결"""
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    """DB 초기화 (통계/분석용)"""
    conn = get_db()
    cur = conn.cursor()

    # 폐기불량 데이터 테이블 (통계/분석용)
    cur.execute('''
        CREATE TABLE IF NOT EXISTS scrap_data (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            unique_id TEXT UNIQUE NOT NULL,
            date TIMESTAMP NOT NULL,
            part TEXT NOT NULL,
            department TEXT,
            process TEXT,
            machine TEXT,
            person TEXT,
            tmno TEXT,
            product_name TEXT,
            scrap_reason TEXT NOT NULL,
            quantity REAL DEFAULT 0,
            weight REAL DEFAULT 0,
            remark TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # 통계용 인덱스 생성
    cur.execute('CREATE INDEX IF NOT EXISTS idx_scrap_date ON scrap_data(date)')
    cur.execute('CREATE INDEX IF NOT EXISTS idx_scrap_part ON scrap_data(part)')
    cur.execute('CREATE INDEX IF NOT EXISTS idx_scrap_department ON scrap_data(department)')
    cur.execute('CREATE INDEX IF NOT EXISTS idx_scrap_process ON scrap_data(process)')
    cur.execute('CREATE INDEX IF NOT EXISTS idx_scrap_reason ON scrap_data(scrap_reason)')
    cur.execute('CREATE INDEX IF NOT EXISTS idx_scrap_tmno ON scrap_data(tmno)')

    conn.commit()
    conn.close()
    print(f"DB 초기화 완료: {DB_FILE}")


def generate_unique_id():
    """고유 ID 생성"""
    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
    random_num = str(random.randint(0, 999)).zfill(3)
    return f'SC{timestamp}{random_num}'


# ==================== 라우트 ====================

@app.route('/')
def index():
    """메인 페이지"""
    return render_template('index.html')


# ==================== API 엔드포인트 (Excel 기반) ====================

@app.route('/api/departments', methods=['GET'])
def get_departments():
    """부서 목록 (Excel)"""
    wb = get_workbook()
    ws = wb['Depart']
    departments = [row[0] for row in ws.iter_rows(min_row=2, values_only=True) if row[0]]
    wb.close()
    return jsonify(departments)


@app.route('/api/processes', methods=['GET'])
def get_processes():
    """공정 목록 (Excel, Part별 필터링)"""
    part = request.args.get('part', '')
    wb = get_workbook()
    ws = wb['Process']

    processes = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            if not part or row[0] == part:
                processes.append(row[1])

    wb.close()
    return jsonify(list(set(processes)))


@app.route('/api/machines', methods=['GET'])
def get_machines():
    """설비 목록 (Excel, Part+공정별 필터링)"""
    part = request.args.get('part', '')
    process = request.args.get('process', '')

    wb = get_workbook()
    ws = wb['machine']

    machines = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1] and row[2]:
            if row[0] == part and row[1] == process:
                machines.append(row[2])

    wb.close()
    return jsonify(machines)


@app.route('/api/persons', methods=['GET'])
def get_persons():
    """폐기자 목록 (Excel, Part+공정+부서별 필터링)"""
    part = request.args.get('part', '')
    process = request.args.get('process', '')
    department = request.args.get('department', '')

    wb = get_workbook()
    ws = wb['person']

    persons = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[3]:
            match = True
            if part and row[0] != part:
                match = False
            if process and row[1] != process:
                match = False
            if department and row[2] != department:
                match = False
            if match:
                persons.append(row[3])

    wb.close()
    return jsonify(persons)


@app.route('/api/tmnos', methods=['GET'])
def get_tmnos():
    """TM-NO 목록 (Excel)"""
    part = request.args.get('part', '1Part')
    process = request.args.get('process', '')

    sheet_name = '1Part_TMNO' if part == '1Part' else '2Part_TMNO'

    if process not in ['성형', '소결']:
        mapped_process = '후처리'
    else:
        mapped_process = process

    wb = get_workbook()
    ws = wb[sheet_name]

    headers = [cell.value for cell in ws[1]]
    process_col = headers.index(mapped_process) if mapped_process in headers else -1

    tmnos = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            if process_col >= 0 and row[process_col]:
                if str(row[process_col]).lower() == 'y':
                    tmnos.append(str(row[0]))
            elif process_col < 0:
                tmnos.append(str(row[0]))

    wb.close()
    return jsonify(tmnos)


@app.route('/api/tmno_info', methods=['GET'])
def get_tmno_info():
    """TM-NO 정보 (Excel)"""
    part = request.args.get('part', '1Part')
    tmno = request.args.get('tmno', '')

    sheet_name = '1Part_TMNO' if part == '1Part' else '2Part_TMNO'

    wb = get_workbook()
    ws = wb[sheet_name]

    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[0]) == str(tmno):
            wb.close()
            return jsonify({
                'tmno': str(row[0]),
                'productName': row[1] or '',
                'unitWeight': float(row[2]) if row[2] else 0
            })

    wb.close()
    return jsonify({'tmno': tmno, 'productName': '', 'unitWeight': 0})


@app.route('/api/scrap_reasons', methods=['GET'])
def get_scrap_reasons():
    """폐기사유 목록 (Excel)"""
    wb = get_workbook()
    ws = wb['scrap_name']
    reasons = [row[0] for row in ws.iter_rows(min_row=2, values_only=True) if row[0]]
    wb.close()
    return jsonify(reasons)


@app.route('/api/save_scrap', methods=['POST'])
def save_scrap():
    """폐기불량 데이터 저장 (Excel + DB 둘 다)"""
    data = request.json

    unique_id = generate_unique_id()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # 1. Excel에 저장
    wb = get_workbook()
    ws = wb['Data']

    new_row = [
        unique_id,
        now,
        data.get('part', ''),
        data.get('department', ''),
        data.get('process', ''),
        data.get('machine', ''),
        data.get('person', ''),
        str(data.get('tmno', '')),
        data.get('productName', ''),
        data.get('scrapReason', ''),
        data.get('quantity', 0),
        data.get('weight', 0),
        data.get('remark', '')
    ]

    ws.append(new_row)
    save_workbook(wb)
    wb.close()

    # 2. DB에 저장 (통계/분석용)
    conn = get_db()
    cur = conn.cursor()

    cur.execute('''
        INSERT INTO scrap_data
        (unique_id, date, part, department, process, machine, person, tmno, product_name, scrap_reason, quantity, weight, remark)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        unique_id,
        now,
        data.get('part', ''),
        data.get('department', ''),
        data.get('process', ''),
        data.get('machine', ''),
        data.get('person', ''),
        str(data.get('tmno', '')),
        data.get('productName', ''),
        data.get('scrapReason', ''),
        data.get('quantity', 0),
        data.get('weight', 0),
        data.get('remark', '')
    ))

    conn.commit()
    conn.close()

    return jsonify({'success': True, 'message': '저장되었습니다.', 'id': unique_id})


# ==================== 마스터 데이터 관리 API (Excel 기반) ====================

@app.route('/api/master_data/<sheet_name>', methods=['GET'])
def get_master_data(sheet_name):
    """마스터 데이터 목록 (Excel)"""
    if sheet_name not in SHEET_STRUCTURES:
        return jsonify({'success': False, 'message': '시트를 찾을 수 없습니다.'})

    wb = get_workbook()
    ws = wb[sheet_name]

    headers = [cell.value for cell in ws[1]]
    rows = []

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if any(row):
            rows.append({
                'rowIndex': idx,
                'data': list(row)
            })

    wb.close()
    return jsonify({'success': True, 'headers': headers, 'rows': rows})


@app.route('/api/master_data/<sheet_name>', methods=['POST'])
def add_master_data(sheet_name):
    """마스터 데이터 추가 (Excel)"""
    if sheet_name not in SHEET_STRUCTURES:
        return jsonify({'success': False, 'message': '시트를 찾을 수 없습니다.'})

    data = request.json.get('data', [])

    wb = get_workbook()
    ws = wb[sheet_name]
    ws.append(data)
    save_workbook(wb)
    wb.close()

    return jsonify({'success': True, 'message': '추가되었습니다.'})


@app.route('/api/master_data/<sheet_name>/<int:row_index>', methods=['PUT'])
def update_master_data(sheet_name, row_index):
    """마스터 데이터 수정 (Excel)"""
    password = request.json.get('password', '')

    if password != ADMIN_PASSWORD:
        return jsonify({'success': False, 'message': '비밀번호가 올바르지 않습니다.'})

    if sheet_name not in SHEET_STRUCTURES:
        return jsonify({'success': False, 'message': '시트를 찾을 수 없습니다.'})

    data = request.json.get('data', [])

    wb = get_workbook()
    ws = wb[sheet_name]

    for col, value in enumerate(data, start=1):
        ws.cell(row=row_index, column=col, value=value)

    save_workbook(wb)
    wb.close()

    return jsonify({'success': True, 'message': '수정되었습니다.'})


@app.route('/api/master_data/<sheet_name>/<int:row_index>', methods=['DELETE'])
def delete_master_data(sheet_name, row_index):
    """마스터 데이터 삭제 (Excel)"""
    password = request.json.get('password', '')

    if password != ADMIN_PASSWORD:
        return jsonify({'success': False, 'message': '비밀번호가 올바르지 않습니다.'})

    if sheet_name not in SHEET_STRUCTURES:
        return jsonify({'success': False, 'message': '시트를 찾을 수 없습니다.'})

    wb = get_workbook()
    ws = wb[sheet_name]
    ws.delete_rows(row_index)
    save_workbook(wb)
    wb.close()

    return jsonify({'success': True, 'message': '삭제되었습니다.'})


@app.route('/api/scrap_records', methods=['GET'])
def get_scrap_records():
    """폐기불량 기록 조회 (Excel)"""
    wb = get_workbook()
    ws = wb['Data']

    headers = [cell.value for cell in ws[1]]
    rows = []

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if any(row):
            rows.append({
                'rowIndex': idx,
                'data': list(row)
            })

    wb.close()
    return jsonify({'success': True, 'headers': headers, 'rows': rows})


@app.route('/api/verify_password', methods=['POST'])
def verify_password():
    """관리자 비밀번호 확인"""
    password = request.json.get('password', '')
    return jsonify(password == ADMIN_PASSWORD)


@app.route('/api/simple_list/<sheet_name>', methods=['GET'])
def get_simple_list(sheet_name):
    """드롭다운용 간단한 목록 (Excel)"""
    if sheet_name not in SHEET_STRUCTURES:
        return jsonify([])

    wb = get_workbook()
    ws = wb[sheet_name]
    items = [row[0] for row in ws.iter_rows(min_row=2, values_only=True) if row[0]]
    wb.close()
    return jsonify(items)


@app.route('/api/process_list', methods=['GET'])
def get_process_list():
    """공정 목록 (Excel)"""
    part = request.args.get('part', '')
    wb = get_workbook()
    ws = wb['Process']

    processes = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1]:
            if not part or row[0] == part:
                processes.append(row[1])

    wb.close()
    return jsonify(list(set(processes)))


# ==================== 통계/분석 API (DB 기반) ====================

@app.route('/api/stats/summary', methods=['GET'])
def get_stats_summary():
    """전체 통계 요약 (DB)"""
    conn = get_db()
    cur = conn.cursor()

    cur.execute('SELECT COUNT(*) as cnt FROM scrap_data')
    total_count = cur.fetchone()['cnt']

    cur.execute('SELECT SUM(quantity) as qty, SUM(weight) as wgt FROM scrap_data')
    row = cur.fetchone()
    total_quantity = row['qty'] or 0
    total_weight = row['wgt'] or 0

    today = datetime.now().strftime('%Y-%m-%d')
    cur.execute("SELECT COUNT(*) as cnt, SUM(quantity) as qty, SUM(weight) as wgt FROM scrap_data WHERE date LIKE ?", (f'{today}%',))
    today_row = cur.fetchone()

    conn.close()

    return jsonify({
        'total': {
            'count': total_count,
            'quantity': total_quantity,
            'weight': round(total_weight, 2)
        },
        'today': {
            'count': today_row['cnt'] or 0,
            'quantity': today_row['qty'] or 0,
            'weight': round(today_row['wgt'] or 0, 2)
        }
    })


@app.route('/api/stats/by_reason', methods=['GET'])
def get_stats_by_reason():
    """폐기사유별 통계 (DB)"""
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')

    conn = get_db()
    cur = conn.cursor()

    query = '''
        SELECT scrap_reason, COUNT(*) as cnt, SUM(quantity) as qty, SUM(weight) as wgt
        FROM scrap_data WHERE 1=1
    '''
    params = []

    if start_date:
        query += ' AND date >= ?'
        params.append(start_date)
    if end_date:
        query += ' AND date <= ?'
        params.append(end_date + ' 23:59:59')

    query += ' GROUP BY scrap_reason ORDER BY cnt DESC'

    cur.execute(query, params)
    results = [{
        'reason': row['scrap_reason'],
        'count': row['cnt'],
        'quantity': row['qty'] or 0,
        'weight': round(row['wgt'] or 0, 2)
    } for row in cur.fetchall()]

    conn.close()
    return jsonify(results)


@app.route('/api/stats/by_part', methods=['GET'])
def get_stats_by_part():
    """Part별 통계 (DB)"""
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')

    conn = get_db()
    cur = conn.cursor()

    query = '''
        SELECT part, COUNT(*) as cnt, SUM(quantity) as qty, SUM(weight) as wgt
        FROM scrap_data WHERE 1=1
    '''
    params = []

    if start_date:
        query += ' AND date >= ?'
        params.append(start_date)
    if end_date:
        query += ' AND date <= ?'
        params.append(end_date + ' 23:59:59')

    query += ' GROUP BY part ORDER BY cnt DESC'

    cur.execute(query, params)
    results = [{
        'part': row['part'],
        'count': row['cnt'],
        'quantity': row['qty'] or 0,
        'weight': round(row['wgt'] or 0, 2)
    } for row in cur.fetchall()]

    conn.close()
    return jsonify(results)


@app.route('/api/stats/by_process', methods=['GET'])
def get_stats_by_process():
    """공정별 통계 (DB)"""
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')

    conn = get_db()
    cur = conn.cursor()

    query = '''
        SELECT process, COUNT(*) as cnt, SUM(quantity) as qty, SUM(weight) as wgt
        FROM scrap_data WHERE 1=1
    '''
    params = []

    if start_date:
        query += ' AND date >= ?'
        params.append(start_date)
    if end_date:
        query += ' AND date <= ?'
        params.append(end_date + ' 23:59:59')

    query += ' GROUP BY process ORDER BY cnt DESC'

    cur.execute(query, params)
    results = [{
        'process': row['process'],
        'count': row['cnt'],
        'quantity': row['qty'] or 0,
        'weight': round(row['wgt'] or 0, 2)
    } for row in cur.fetchall()]

    conn.close()
    return jsonify(results)


@app.route('/api/stats/daily', methods=['GET'])
def get_stats_daily():
    """일별 통계 (DB)"""
    days = request.args.get('days', 30, type=int)

    conn = get_db()
    cur = conn.cursor()

    cur.execute('''
        SELECT DATE(date) as day, COUNT(*) as cnt, SUM(quantity) as qty, SUM(weight) as wgt
        FROM scrap_data
        WHERE date >= DATE('now', ? || ' days')
        GROUP BY DATE(date)
        ORDER BY day DESC
    ''', (f'-{days}',))

    results = [{
        'date': row['day'],
        'count': row['cnt'],
        'quantity': row['qty'] or 0,
        'weight': round(row['wgt'] or 0, 2)
    } for row in cur.fetchall()]

    conn.close()
    return jsonify(results)


if __name__ == '__main__':
    init_excel_file()
    init_db()
    print("=" * 50)
    print("폐기불량 관리시스템 서버 시작")
    print("http://localhost:5001 에서 접속하세요")
    print(f"Excel 파일: {EXCEL_FILE}")
    print(f"SQLite DB (통계용): {DB_FILE}")
    print("=" * 50)
    app.run(host='0.0.0.0', port=5001, debug=True)
